import io
import math
from collections import deque

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DIAS = [
    "Lunes",
    "Martes",
    "Miércoles",
    "Jueves",
    "Viernes",
    "Sábado",
    "Domingo",
]

# Es el ciclo que aparece en el Excel de ejemplo.
CICLO_12X24 = ["DIA", "NOCHE", "DESCANSO"]
LOGO_URL = "https://seguridadtotal.mx/wp-content/uploads/2026/04/STALOGO-1.png"


class Dinic:
    """Flujo máximo pequeño para repartir oficiales sin dejar huecos."""

    def __init__(self, cantidad_nodos: int):
        self.grafo = [[] for _ in range(cantidad_nodos)]

    def agregar_arista(self, origen: int, destino: int, capacidad: int):
        directa = [destino, capacidad, None]
        inversa = [origen, 0, directa]
        directa[2] = inversa
        self.grafo[origen].append(directa)
        self.grafo[destino].append(inversa)
        return directa

    def flujo_maximo(self, origen: int, destino: int) -> int:
        flujo_total = 0

        while True:
            nivel = [-1] * len(self.grafo)
            nivel[origen] = 0
            cola = deque([origen])

            while cola:
                nodo = cola.popleft()
                for arista in self.grafo[nodo]:
                    siguiente, capacidad, _ = arista
                    if capacidad > 0 and nivel[siguiente] < 0:
                        nivel[siguiente] = nivel[nodo] + 1
                        cola.append(siguiente)

            if nivel[destino] < 0:
                return flujo_total

            posicion = [0] * len(self.grafo)

            def enviar(nodo: int, disponible: int) -> int:
                if nodo == destino:
                    return disponible

                while posicion[nodo] < len(self.grafo[nodo]):
                    arista = self.grafo[nodo][posicion[nodo]]
                    siguiente, capacidad, inversa = arista

                    if capacidad > 0 and nivel[siguiente] == nivel[nodo] + 1:
                        enviado = enviar(siguiente, min(disponible, capacidad))
                        if enviado:
                            arista[1] -= enviado
                            inversa[1] += enviado
                            return enviado

                    posicion[nodo] += 1

                return 0

            while True:
                enviado = enviar(origen, 10**9)
                if enviado == 0:
                    break
                flujo_total += enviado


def calcular_horas(turnos: list[str], horas_turno: int) -> int:
    return sum(turno in ("DIA", "NOCHE") for turno in turnos) * horas_turno


def cumple_descansos(turnos: list[str], max_descansos: int = 2) -> bool:
    consecutivos = 0

    for turno in turnos:
        if turno == "DESCANSO":
            consecutivos += 1

            if consecutivos > max_descansos:
                return False
        else:
            consecutivos = 0

    return True


def generar_patrones_labor(max_turnos_guardia: int) -> list[tuple[bool, ...]]:
    patrones = []

    def construir(
        dia: int,
        actual: list[bool],
        labores: int,
        descansos_seguidos: int,
    ):
        if dia == 7:
            patron_14 = (actual * 2)

            if any(actual) and cumple_descansos(
                ["DIA" if trabaja else "DESCANSO" for trabaja in patron_14]
            ):
                patrones.append(tuple(actual))
            return

        for trabaja in (True, False):
            nuevo_descanso = 0 if trabaja else descansos_seguidos + 1

            if nuevo_descanso > 2:
                continue

            nuevas_labores = labores + int(trabaja)

            if nuevas_labores > max_turnos_guardia:
                continue

            actual.append(trabaja)
            construir(
                dia + 1,
                actual,
                nuevas_labores,
                nuevo_descanso,
            )
            actual.pop()

    construir(0, [], 0, 0)
    patrones.sort(key=sum, reverse=True)

    return patrones


def crear_roles_por_patrones(
    necesidades_dia: list[int],
    necesidades_noche: list[int],
    permite_dia: bool,
    permite_noche: bool,
    nombre_grupo: str,
    nombre_puesto: str,
    max_turnos_guardia: int,
    numero_inicial: int,
) -> tuple[list[dict], int]:
    necesidades_por_dia = tuple(
        necesidades_dia[dia] + necesidades_noche[dia]
        for dia in range(7)
    )
    total_turnos = sum(necesidades_por_dia)

    if total_turnos == 0:
        return [], numero_inicial

    for dia, cantidad in enumerate(necesidades_por_dia):
        cobertura_previa = sum(
            necesidades_por_dia[(dia - distancia) % 7]
            for distancia in range(1, 4)
        )

        if cantidad > cobertura_previa:
            raise ValueError(
                "No fue posible generar el rol con máximo 2 descansos seguidos. "
                f"Para cubrir {DIAS[dia]} con {cantidad} oficiales, necesitas "
                "más labores en los tres días anteriores o bajar esa cobertura."
            )

    patrones = generar_patrones_labor(max_turnos_guardia)

    minimo_guardias = max(
        max(necesidades_por_dia),
        math.ceil(total_turnos / max_turnos_guardia),
        1,
    )

    def cabe(patron: tuple[bool, ...], restantes: tuple[int, ...]) -> bool:
        return all(int(trabaja) <= restantes[dia] for dia, trabaja in enumerate(patron))

    def restar(
        patron: tuple[bool, ...],
        restantes: tuple[int, ...],
    ) -> tuple[int, ...]:
        return tuple(restantes[dia] - int(patron[dia]) for dia in range(7))

    def buscar(
        restantes: tuple[int, ...],
        guardias_disponibles: int,
        memoria: set[tuple[tuple[int, ...], int]],
    ) -> list[tuple[str, ...]] | None:
        if sum(restantes) == 0:
            return []

        if guardias_disponibles == 0:
            return None

        estado = (restantes, guardias_disponibles)
        if estado in memoria:
            return None

        if max(restantes) > guardias_disponibles:
            memoria.add(estado)
            return None

        if sum(restantes) > guardias_disponibles * max_turnos_guardia:
            memoria.add(estado)
            return None

        dia_objetivo = max(range(7), key=lambda dia: restantes[dia])

        for patron in patrones:
            if not patron[dia_objetivo]:
                continue

            if not cabe(patron, restantes):
                continue

            solucion = buscar(
                restar(patron, restantes),
                guardias_disponibles - 1,
                memoria,
            )

            if solucion is not None:
                return [patron] + solucion

        memoria.add(estado)
        return None

    for numero_guardias in range(minimo_guardias, total_turnos + 1):
        solucion = buscar(necesidades_por_dia, numero_guardias, set())

        if solucion is None:
            continue

        roles = []
        numero_oficial = numero_inicial

        turnos_por_rol = [["DESCANSO"] * 14 for _ in solucion]

        for dia_semana in range(7):
            trabajadores = [
                indice
                for indice, patron in enumerate(solucion)
                if patron[dia_semana]
            ]
            dia_requeridos = necesidades_dia[dia_semana]
            noche_requeridos = necesidades_noche[dia_semana]

            if not permite_dia:
                dia_requeridos = 0

            if not permite_noche:
                noche_requeridos = 0

            for indice in trabajadores[:dia_requeridos]:
                turnos_por_rol[indice][dia_semana] = "DIA"

            for indice in trabajadores[dia_requeridos:dia_requeridos + noche_requeridos]:
                turnos_por_rol[indice][dia_semana] = "NOCHE"

        rotacion_semana_2 = None

        for desplazamiento in range(1, len(solucion)):
            candidato = [turnos[:] for turnos in turnos_por_rol]

            for indice in range(len(candidato)):
                origen = (indice + desplazamiento) % len(candidato)
                candidato[indice][7:] = turnos_por_rol[origen][:7]

            if all(cumple_descansos(turnos) for turnos in candidato):
                rotacion_semana_2 = candidato
                break

        if rotacion_semana_2 is None:
            rotacion_semana_2 = [turnos[:] for turnos in turnos_por_rol]

            for indice in range(len(rotacion_semana_2)):
                rotacion_semana_2[indice][7:] = turnos_por_rol[indice][:7]

        turnos_por_rol = rotacion_semana_2

        for indice, turnos in enumerate(turnos_por_rol):
            roles.append(
                {
                    "grupo": nombre_grupo,
                    "puesto": f"OFICIAL {indice + 1}",
                    "numero": numero_oficial,
                    "turnos": turnos,
                }
            )
            numero_oficial += 1

        return roles, numero_oficial

    raise ValueError(
        "No fue posible generar un rol con máximo 2 descansos seguidos."
    )


def asignar_semana_fija(
    necesidades: list[int],
    numero_guardias: int,
    max_turnos_guardia: int,
    desplazamiento: int,
) -> list[list[bool]]:
    """Crea una matriz guardia x día que cumple exactamente la cobertura."""

    total_turnos = sum(necesidades)

    if numero_guardias == 0:
        return []

    if total_turnos == 0:
        return [[False] * 7 for _ in range(numero_guardias)]

    turnos_base, extras = divmod(total_turnos, numero_guardias)
    objetivos = [turnos_base] * numero_guardias

    for indice in range(extras):
        objetivos[(desplazamiento + indice) % numero_guardias] += 1

    if max(objetivos) > max_turnos_guardia:
        raise ValueError(
            "No es posible repartir la cobertura con el máximo de turnos indicado."
        )

    fuente = 0
    primer_guardia = 1
    primer_dia = primer_guardia + numero_guardias
    sumidero = primer_dia + 7

    flujo = Dinic(sumidero + 1)
    referencias = {}

    # El orden se desplaza en la segunda semana para que no siempre descansen
    # las mismas personas en los mismos días.
    orden_guardias = [
        (desplazamiento + indice) % numero_guardias
        for indice in range(numero_guardias)
    ]

    for guardia in orden_guardias:
        flujo.agregar_arista(
            fuente,
            primer_guardia + guardia,
            objetivos[guardia],
        )

        orden_dias = sorted(
            range(7),
            key=lambda dia: (dia - guardia - desplazamiento) % 7,
        )

        for dia in orden_dias:
            referencias[(guardia, dia)] = flujo.agregar_arista(
                primer_guardia + guardia,
                primer_dia + dia,
                1,
            )

    for dia, cantidad in enumerate(necesidades):
        flujo.agregar_arista(primer_dia + dia, sumidero, cantidad)

    obtenido = flujo.flujo_maximo(fuente, sumidero)

    if obtenido != total_turnos:
        raise ValueError(
            "No se pudo construir un rol que cubra todos los turnos. "
            "Prueba aumentando el máximo de turnos por guardia."
        )

    matriz = [[False] * 7 for _ in range(numero_guardias)]

    for (guardia, dia), arista in referencias.items():
        # Una arista con capacidad restante 0 fue utilizada por el flujo.
        if arista[1] == 0:
            matriz[guardia][dia] = True

    return matriz


def asignar_semana_mixta(
    necesidades_dia: list[int],
    necesidades_noche: list[int],
    numero_guardias: int,
    max_turnos_guardia: int,
    desplazamiento: int,
) -> list[list[str]]:
    total_turnos = sum(necesidades_dia) + sum(necesidades_noche)

    if numero_guardias == 0:
        return []

    if total_turnos == 0:
        return [["DESCANSO"] * 7 for _ in range(numero_guardias)]

    turnos_base, extras = divmod(total_turnos, numero_guardias)
    objetivos = [turnos_base] * numero_guardias

    for indice in range(extras):
        objetivos[(desplazamiento + indice) % numero_guardias] += 1

    if max(objetivos) > max_turnos_guardia:
        raise ValueError(
            "No es posible repartir la cobertura con el máximo de turnos indicado."
        )

    fuente = 0
    primer_guardia = 1
    primer_guardia_dia = primer_guardia + numero_guardias
    primer_turno = primer_guardia_dia + numero_guardias * 7
    sumidero = primer_turno + 14

    flujo = Dinic(sumidero + 1)
    referencias = {}

    orden_guardias = [
        (desplazamiento + indice) % numero_guardias
        for indice in range(numero_guardias)
    ]

    for guardia in orden_guardias:
        flujo.agregar_arista(
            fuente,
            primer_guardia + guardia,
            objetivos[guardia],
        )

        orden_dias = sorted(
            range(7),
            key=lambda dia: (dia - guardia - desplazamiento) % 7,
        )

        for dia in orden_dias:
            nodo_guardia_dia = primer_guardia_dia + guardia * 7 + dia
            flujo.agregar_arista(
                primer_guardia + guardia,
                nodo_guardia_dia,
                1,
            )

            for tipo, offset in (("DIA", 0), ("NOCHE", 7)):
                referencias[(guardia, dia, tipo)] = flujo.agregar_arista(
                    nodo_guardia_dia,
                    primer_turno + offset + dia,
                    1,
                )

    for dia, cantidad in enumerate(necesidades_dia):
        flujo.agregar_arista(primer_turno + dia, sumidero, cantidad)

    for dia, cantidad in enumerate(necesidades_noche):
        flujo.agregar_arista(primer_turno + 7 + dia, sumidero, cantidad)

    obtenido = flujo.flujo_maximo(fuente, sumidero)

    if obtenido != total_turnos:
        raise ValueError(
            "No se pudo construir un rol que cubra todos los turnos. "
            "Prueba revisando la cobertura o los descansos."
        )

    matriz = [["DESCANSO"] * 7 for _ in range(numero_guardias)]

    for (guardia, dia, tipo), arista in referencias.items():
        if arista[1] == 0:
            matriz[guardia][dia] = tipo

    return matriz


def crear_oficiales_mixtos(
    necesidades_dia: list[int],
    necesidades_noche: list[int],
    nombre_grupo: str,
    nombre_puesto: str,
    max_turnos_guardia: int,
    numero_inicial: int,
) -> tuple[list[dict], int]:
    return crear_roles_por_patrones(
        necesidades_dia=necesidades_dia,
        necesidades_noche=necesidades_noche,
        permite_dia=True,
        permite_noche=True,
        nombre_grupo=nombre_grupo,
        nombre_puesto=nombre_puesto,
        max_turnos_guardia=max_turnos_guardia,
        numero_inicial=numero_inicial,
    )


def crear_oficiales_fijos(
    necesidades: list[int],
    tipo_turno: str,
    nombre_grupo: str,
    nombre_puesto: str,
    max_turnos_guardia: int,
    numero_inicial: int,
) -> tuple[list[dict], int]:
    return crear_roles_por_patrones(
        necesidades_dia=necesidades if tipo_turno == "DIA" else [0] * 7,
        necesidades_noche=necesidades if tipo_turno == "NOCHE" else [0] * 7,
        permite_dia=tipo_turno == "DIA",
        permite_noche=tipo_turno == "NOCHE",
        nombre_grupo=nombre_grupo,
        nombre_puesto=nombre_puesto,
        max_turnos_guardia=max_turnos_guardia,
        numero_inicial=numero_inicial,
    )


def crear_oficiales_dia_fijos(
    necesidades: list[int],
    numero_inicial: int,
) -> tuple[list[dict], list[int], int]:
    dias_fijos = 5
    cantidad_fija = min(necesidades[:dias_fijos], default=0)

    roles = []
    numero_oficial = numero_inicial

    for indice in range(cantidad_fija):
        semana = [
            "DIA" if dia < dias_fijos and indice < necesidades[dia] else "DESCANSO"
            for dia in range(7)
        ]

        roles.append(
            {
                "grupo": "GRUPO 1",
                "puesto": f"OFICIAL {indice + 1}",
                "numero": numero_oficial,
                "turnos": semana + semana,
            }
        )
        numero_oficial += 1

    faltantes = [
        max(0, cantidad - cantidad_fija) if dia < dias_fijos else cantidad
        for dia, cantidad in enumerate(necesidades)
    ]

    return roles, faltantes, numero_oficial


def crear_grupos_mixtos_espejo(
    necesidades_dia: list[int],
    necesidades_noche: list[int],
    numero_inicial: int,
) -> tuple[list[dict], int] | None:
    if any(necesidades_dia[dia] for dia in range(5)):
        return None

    if necesidades_dia[5] != necesidades_dia[6]:
        return None

    if len(set(necesidades_noche)) != 1:
        return None

    cantidad_grupo = necesidades_noche[0]

    if cantidad_grupo == 0:
        return [], numero_inicial

    if necesidades_dia[5] != cantidad_grupo:
        return None

    patron_grupo_2 = [
        "NOCHE",
        "DESCANSO",
        "NOCHE",
        "DESCANSO",
        "NOCHE",
        "NOCHE",
        "NOCHE",
    ]
    patron_grupo_3 = [
        "DESCANSO",
        "NOCHE",
        "DESCANSO",
        "NOCHE",
        "DESCANSO",
        "DIA",
        "DIA",
    ]

    roles = []
    numero_oficial = numero_inicial

    for indice in range(cantidad_grupo):
        roles.append(
            {
                "grupo": "GRUPO 2",
                "puesto": f"OFICIAL {indice + 1}",
                "numero": numero_oficial,
                "turnos": patron_grupo_2 + patron_grupo_3,
            }
        )
        numero_oficial += 1

    for indice in range(cantidad_grupo):
        roles.append(
            {
                "grupo": "GRUPO 3",
                "puesto": f"OFICIAL {indice + 1}",
                "numero": numero_oficial,
                "turnos": patron_grupo_3 + patron_grupo_2,
            }
        )
        numero_oficial += 1

    return roles, numero_oficial


def agrupar_roles_por_patron(roles: list[dict]) -> list[dict]:
    if not roles:
        return []

    cantidad_grupos = min(3, len(roles))
    base, extras = divmod(len(roles), cantidad_grupos)
    roles_agrupados = []
    numero_oficial = 1
    inicio = 0

    for indice_grupo in range(cantidad_grupos):
        tamano_grupo = base + int(indice_grupo < extras)
        bloque = roles[inicio:inicio + tamano_grupo]

        for rol in bloque:
            rol["grupo"] = f"GRUPO {indice_grupo + 1}"
            rol["puesto"] = f"OFICIAL {numero_oficial}"
            rol["numero"] = numero_oficial
            roles_agrupados.append(rol)
            numero_oficial += 1

        inicio += tamano_grupo

    return roles_agrupados


def generar_roles(
    guardias_dia: list[int],
    guardias_noche: list[int],
    posiciones_12x24: int,
    max_turnos_guardia: int,
    dia_fijo: bool,
) -> list[dict]:
    maximo_12x24 = min(min(guardias_dia), min(guardias_noche))

    if posiciones_12x24 > maximo_12x24:
        raise ValueError(
            "Las posiciones 12x24 no pueden superar la cobertura mínima "
            "simultánea de día y noche."
        )

    roles = []
    numero_oficial = 1

    # Cada posición 12x24 necesita tres personas desfasadas:
    # Grupo 1: DIA, NOCHE, DESCANSO...
    # Grupo 2: NOCHE, DESCANSO, DIA...
    # Grupo 3: DESCANSO, DIA, NOCHE...
    for grupo in range(3):
        for posicion in range(posiciones_12x24):
            turnos = [
                CICLO_12X24[(dia + grupo) % len(CICLO_12X24)]
                for dia in range(14)
            ]

            roles.append(
                {
                    "grupo": f"GRUPO {grupo + 1}",
                    "puesto": f"OFICIAL {posicion + 1}",
                    "numero": numero_oficial,
                    "turnos": turnos,
                }
            )
            numero_oficial += 1

    faltantes_dia = [
        cantidad - posiciones_12x24 for cantidad in guardias_dia
    ]
    faltantes_noche = [
        cantidad - posiciones_12x24 for cantidad in guardias_noche
    ]

    if dia_fijo:
        oficiales_fijos, faltantes_dia, numero_oficial = crear_oficiales_dia_fijos(
            faltantes_dia,
            numero_oficial,
        )
        roles.extend(oficiales_fijos)

        grupos_espejo = crear_grupos_mixtos_espejo(
            faltantes_dia,
            faltantes_noche,
            numero_oficial,
        )

        if grupos_espejo is not None:
            oficiales_mixtos, numero_oficial = grupos_espejo
            roles.extend(oficiales_mixtos)
            return agrupar_roles_por_patron(roles)

        oficiales_mixtos, numero_oficial = crear_oficiales_mixtos(
            necesidades_dia=faltantes_dia,
            necesidades_noche=faltantes_noche,
            nombre_grupo="GRUPO 2",
            nombre_puesto="OFICIAL MIXTO",
            max_turnos_guardia=max_turnos_guardia,
            numero_inicial=numero_oficial,
        )
        roles.extend(oficiales_mixtos)

        return agrupar_roles_por_patron(roles)

    oficiales_dia, numero_oficial = crear_oficiales_fijos(
        necesidades=faltantes_dia,
        tipo_turno="DIA",
        nombre_grupo="TURNO DIA",
        nombre_puesto="OFICIAL DIA",
        max_turnos_guardia=max_turnos_guardia,
        numero_inicial=numero_oficial,
    )
    roles.extend(oficiales_dia)

    oficiales_noche, numero_oficial = crear_oficiales_fijos(
        necesidades=faltantes_noche,
        tipo_turno="NOCHE",
        nombre_grupo="TURNO NOCHE",
        nombre_puesto="OFICIAL NOCHE",
        max_turnos_guardia=max_turnos_guardia,
        numero_inicial=numero_oficial,
    )
    roles.extend(oficiales_noche)

    return agrupar_roles_por_patron(roles)


def contar_cobertura(
    roles: list[dict],
    inicio_semana: int,
) -> tuple[list[int], list[int]]:
    cobertura_dia = []
    cobertura_noche = []

    for dia in range(7):
        indice = inicio_semana + dia
        cobertura_dia.append(
            sum(rol["turnos"][indice] == "DIA" for rol in roles)
        )
        cobertura_noche.append(
            sum(rol["turnos"][indice] == "NOCHE" for rol in roles)
        )

    return cobertura_dia, cobertura_noche


def crear_tabla_previa(
    roles: list[dict],
    horas_turno: int,
    horas_objetivo: int,
) -> pd.DataFrame:
    filas = []

    for rol in roles:
        semana_1 = rol["turnos"][:7]
        semana_2 = rol["turnos"][7:]
        horas_1 = calcular_horas(semana_1, horas_turno)
        horas_2 = calcular_horas(semana_2, horas_turno)

        fila = {
            "GRUPO": rol["grupo"],
            "PUESTO": rol["puesto"],
            "# OF.": rol["numero"],
        }

        for indice, dia in enumerate(DIAS):
            fila[f"{dia} S1"] = semana_1[indice]

        fila["HRS S1"] = horas_1
        fila["DIF S1"] = horas_1 - horas_objetivo

        for indice, dia in enumerate(DIAS):
            fila[f"{dia} S2"] = semana_2[indice]

        fila["HRS S2"] = horas_2
        fila["DIF S2"] = horas_2 - horas_objetivo
        filas.append(fila)

    return pd.DataFrame(filas)


def generar_excel(
    cliente: str,
    tipo_turnos: str,
    guardias_dia: list[int],
    guardias_noche: list[int],
    roles: list[dict],
    horas_turno: int,
    horas_objetivo: int,
) -> bytes:
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Roles"

    verde = PatternFill("solid", fgColor="008000")
    azul = PatternFill("solid", fgColor="00CCFF")
    amarillo = PatternFill("solid", fgColor="FFFF00")
    gris = PatternFill("solid", fgColor="D9E1F2")
    gris_oscuro = PatternFill("solid", fgColor="A6A6A6")
    blanco = PatternFill("solid", fgColor="FFFFFF")

    borde_fino = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    hoja.merge_cells("A1:U1")
    hoja["A1"] = f"DETALLADO ROLES DE TURNO — {cliente.upper()}"
    hoja["A1"].font = Font(bold=True, size=15)
    hoja["A1"].alignment = Alignment(horizontal="center")

    hoja.merge_cells("A2:U2")
    hoja["A2"] = f"TIPO DE TURNOS: {tipo_turnos.upper()}"
    hoja["A2"].font = Font(bold=True, size=11)
    hoja["A2"].alignment = Alignment(horizontal="center")

    hoja.merge_cells("A3:U3")
    hoja["A3"] = f"Horas objetivo por semana: {horas_objetivo}."
    hoja["A3"].alignment = Alignment(horizontal="center")

    encabezados = (
        ["GRUPO", "PUESTO", "# OF."]
        + [dia.upper() for dia in DIAS]
        + ["HRS SEM", "DIF. SEM"]
        + [dia.upper() for dia in DIAS]
        + ["HRS SEM", "DIF. SEM"]
    )

    fila_encabezado = 5
    for columna, encabezado in enumerate(encabezados, start=1):
        celda = hoja.cell(fila_encabezado, columna, encabezado)
        celda.fill = gris_oscuro
        celda.font = Font(bold=True, color="FFFFFF")
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = borde_fino

    fila = fila_encabezado + 1
    rangos_grupo = []
    grupo_actual = None
    fila_inicio_grupo = fila

    for rol in roles:
        if grupo_actual is None:
            grupo_actual = rol["grupo"]
            fila_inicio_grupo = fila
        elif rol["grupo"] != grupo_actual:
            rangos_grupo.append((fila_inicio_grupo, fila - 1))
            grupo_actual = rol["grupo"]
            fila_inicio_grupo = fila

        semana_1 = rol["turnos"][:7]
        semana_2 = rol["turnos"][7:]
        horas_1 = calcular_horas(semana_1, horas_turno)
        horas_2 = calcular_horas(semana_2, horas_turno)

        valores = (
            [rol["grupo"], rol["puesto"], rol["numero"]]
            + semana_1
            + [horas_1, horas_1 - horas_objetivo]
            + semana_2
            + [horas_2, horas_2 - horas_objetivo]
        )

        for columna, valor in enumerate(valores, start=1):
            celda = hoja.cell(fila, columna, valor)
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border = borde_fino

            if valor == "DIA":
                celda.fill = verde
            elif valor == "NOCHE":
                celda.fill = azul
            elif valor == "DESCANSO":
                celda.fill = amarillo
            elif columna <= 3:
                celda.fill = gris
            else:
                celda.fill = blanco

        fila += 1

    if grupo_actual is not None:
        rangos_grupo.append((fila_inicio_grupo, fila - 1))

    for inicio_grupo, fin_grupo in rangos_grupo:
        celda_grupo = hoja.cell(inicio_grupo, 1)
        celda_grupo.alignment = Alignment(horizontal="center", vertical="center")
        celda_grupo.font = Font(bold=True)

        if fin_grupo > inicio_grupo:
            hoja.merge_cells(
                start_row=inicio_grupo,
                start_column=1,
                end_row=fin_grupo,
                end_column=1,
            )

    hoja.freeze_panes = "D6"
    hoja.auto_filter.ref = f"A5:U{max(fila - 1, 5)}"
    hoja.sheet_view.showGridLines = False

    anchos = {
        "A": 19,
        "B": 25,
        "C": 8,
        "K": 10,
        "L": 10,
        "T": 10,
        "U": 10,
    }

    for columna in range(1, 22):
        letra = get_column_letter(columna)
        hoja.column_dimensions[letra].width = anchos.get(letra, 14)

    # Segunda hoja: comprobación de cobertura.
    cobertura = libro.create_sheet("Cobertura")
    cobertura.sheet_view.showGridLines = False

    cobertura.merge_cells("A1:I1")
    cobertura["A1"] = f"VALIDACIÓN DE COBERTURA — {cliente.upper()}"
    cobertura["A1"].font = Font(bold=True, size=14)
    cobertura["A1"].alignment = Alignment(horizontal="center")

    encabezado_cobertura = ["CONCEPTO"] + [dia.upper() for dia in DIAS] + ["TOTAL"]
    for columna, valor in enumerate(encabezado_cobertura, start=1):
        celda = cobertura.cell(3, columna, valor)
        celda.fill = gris_oscuro
        celda.font = Font(bold=True, color="FFFFFF")
        celda.alignment = Alignment(horizontal="center")
        celda.border = borde_fino

    dia_s1, noche_s1 = contar_cobertura(roles, 0)
    dia_s2, noche_s2 = contar_cobertura(roles, 7)

    filas_cobertura = [
        ("DIA REQUERIDO", guardias_dia, verde),
        ("DIA GENERADO S1", dia_s1, verde),
        ("DIA GENERADO S2", dia_s2, verde),
        ("NOCHE REQUERIDA", guardias_noche, azul),
        ("NOCHE GENERADA S1", noche_s1, azul),
        ("NOCHE GENERADA S2", noche_s2, azul),
    ]

    for numero_fila, (titulo, datos, relleno) in enumerate(
        filas_cobertura,
        start=4,
    ):
        valores = [titulo] + datos + [sum(datos)]
        for columna, valor in enumerate(valores, start=1):
            celda = cobertura.cell(numero_fila, columna, valor)
            celda.border = borde_fino
            celda.alignment = Alignment(horizontal="center")
            celda.fill = relleno if columna > 1 else gris
            if columna == 1:
                celda.font = Font(bold=True)

    cobertura["A11"] = "RESULTADO"
    cobertura["A11"].font = Font(bold=True)

    cobertura_correcta = (
        dia_s1 == guardias_dia
        and dia_s2 == guardias_dia
        and noche_s1 == guardias_noche
        and noche_s2 == guardias_noche
    )

    cobertura["B11"] = "COBERTURA COMPLETA" if cobertura_correcta else "REVISAR"
    cobertura["B11"].font = Font(bold=True)
    cobertura["B11"].fill = verde if cobertura_correcta else amarillo

    cobertura.column_dimensions["A"].width = 25
    for columna in "BCDEFGHI":
        cobertura.column_dimensions[columna].width = 14

    archivo = io.BytesIO()
    libro.save(archivo)
    archivo.seek(0)
    return archivo.getvalue()

def aplicar_estilos_marca():
    st.markdown(
        """
        <style>
        :root {
            --st-red: #e1261c;
            --st-red-dark: #c90000;
            --st-red-deep: #9c0000;
            --st-black: #050505;
            --st-ink: #111111;
            --st-muted: #5f6368;
            --st-line: #e6e8ec;
            --st-soft: #f7f8fa;
        }

        .stApp {
            background: #ffffff;
            color: var(--st-ink);
        }

        .block-container {
            padding-top: .35rem;
            padding-bottom: 2.75rem;
            max-width: 1380px;
        }

        .st-brand-shell {
            position: relative;
            overflow: hidden;
            min-height: 448px;
            border-radius: 0 0 28px 28px;
            background: #ffffff;
            box-shadow: 0 18px 34px rgba(0, 0, 0, .11);
            animation: stFadeIn .65s ease-out both;
        }

        .st-hero-stage {
            position: relative;
            min-height: 448px;
            display: grid;
            place-items: center;
            padding: 66px 24px 150px;
            background:
                linear-gradient(125deg, rgba(225, 38, 28, .08) 0 22%, transparent 22% 100%),
                linear-gradient(305deg, rgba(0, 0, 0, .045) 0 18%, transparent 18% 100%);
        }

        .st-hero-mark {
            position: absolute;
            left: 6%;
            top: 22%;
            width: clamp(130px, 18vw, 240px);
            aspect-ratio: 1;
            border-radius: 999px;
            border: 18px solid rgba(225, 38, 28, .12);
            box-shadow:
                inset 0 0 0 24px rgba(0, 0, 0, .04),
                0 26px 60px rgba(225, 38, 28, .13);
            z-index: 2;
        }

        .st-hero-mark::after {
            content: "";
            position: absolute;
            inset: 28%;
            border-radius: inherit;
            background: var(--st-red);
        }

        .st-brand-title {
            position: relative;
            z-index: 3;
            margin: 0 0 18px;
            color: var(--st-red-dark);
            font-size: clamp(2.5rem, 5.4vw, 5rem);
            line-height: 1;
            font-weight: 900;
            letter-spacing: 0;
            text-align: center;
        }

        .st-brand-copy {
            position: relative;
            z-index: 3;
            max-width: 780px;
            color: #000000;
            font-size: 1.08rem;
            line-height: 1.55;
            margin: 0 auto;
            text-align: center;
        }

        .st-chip-row {
            display: none;
        }

        .st-hero-content {
            position: relative;
            z-index: 3;
            width: min(760px, 92%);
            margin-left: min(18vw, 180px);
        }

        .st-hero-cta {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            min-height: 52px;
            margin-top: 32px;
            padding: 0 34px;
            border-radius: 10px;
            color: #ffffff;
            background: #000000;
            box-shadow: 7px 8px 0 var(--st-red);
            font-weight: 800;
        }

        .st-wave-black,
        .st-wave-red,
        .st-wave-red-soft {
            position: absolute;
            pointer-events: none;
        }

        .st-wave-black {
            right: -11%;
            bottom: -13%;
            width: 62%;
            height: 58%;
            z-index: 1;
            background: #000000;
            border-radius: 100% 0 0 0;
            transform: rotate(-6deg);
        }

        .st-wave-red {
            right: -12%;
            bottom: -28%;
            width: 88%;
            height: 43%;
            z-index: 2;
            background: var(--st-red-dark);
            border-radius: 100% 0 0 0;
            transform: rotate(-4deg);
        }

        .st-wave-red-soft {
            right: -10%;
            bottom: -26%;
            width: 66%;
            height: 32%;
            z-index: 2;
            background: rgba(225, 38, 28, .42);
            border-radius: 100% 0 0 0;
            transform: rotate(-7deg);
        }

        div[data-testid="stMetric"] {
            background: #ffffff;
            border: 1px solid var(--st-line);
            border-left: 5px solid var(--st-red);
            border-radius: 8px;
            padding: 16px 18px;
            box-shadow: 0 12px 30px rgba(18, 20, 24, .07);
        }

        div[data-testid="stMetricValue"] {
            color: #000000;
            font-weight: 850;
        }

        div[data-testid="stMetricLabel"] {
            color: var(--st-muted);
            font-weight: 750;
        }

        .stButton > button,
        .stDownloadButton > button {
            border-radius: 10px;
            border: 0;
            min-height: 46px;
            font-weight: 800;
            color: #ffffff;
            background: linear-gradient(135deg, var(--st-red) 0%, var(--st-red-dark) 100%);
            box-shadow: 0 12px 24px rgba(195, 23, 24, .25);
        }

        .stButton > button:hover,
        .stDownloadButton > button:hover {
            color: #ffffff;
            transform: translateY(-1px);
            box-shadow: 0 18px 32px rgba(195, 23, 24, .32);
        }

        div[data-testid="stDataFrame"] {
            border: 1px solid var(--st-line);
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 16px 36px rgba(18, 20, 24, .08);
        }

        h2, h3 {
            color: var(--st-red-dark);
            letter-spacing: 0;
        }

        label, .stMarkdown, .stCaption, div[data-testid="stWidgetLabel"] {
            color: #111111 !important;
        }

        input, textarea, select, div[data-baseweb="select"] > div {
            color: #111111 !important;
            background-color: #ffffff !important;
        }

        hr {
            border-color: rgba(36, 39, 45, .12);
        }

        @keyframes stFadeIn {
            from { opacity: 0; transform: translateY(12px); }
            to { opacity: 1; transform: translateY(0); }
        }

        @media (max-width: 760px) {
            .st-hero-stage {
                min-height: 420px;
                padding: 42px 18px 150px;
            }

            .st-hero-content {
                margin-left: 0;
            }

            .st-hero-mark {
                left: -54px;
                top: 56%;
                opacity: .35;
            }

            .st-wave-black {
                width: 105%;
                height: 34%;
                right: -38%;
            }

            .st-wave-red {
                width: 120%;
                height: 28%;
                right: -48%;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_hero():
    st.markdown(
        f"""
        <div class="st-brand-shell">
            <div class="st-hero-stage">
                <div class="st-hero-mark"></div>
                <div class="st-hero-content">
                    <h1 class="st-brand-title">Seguridad Total</h1>
                    <p class="st-brand-copy">
                        Generador de roles para cobertura operativa, supervisión
                        de turnos y exportación profesional para cada servicio.
                    </p>
                    <div class="st-hero-cta">Generador de roles</div>
                </div>
                <div class="st-wave-black"></div>
                <div class="st-wave-red"></div>
                <div class="st-wave-red-soft"></div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def pintar_turnos(tabla: pd.DataFrame):
    def color_celda(valor):
        if valor == "DIA":
            return "background-color: #e7f7ef; color: #0f7a48; font-weight: 700;"
        if valor == "NOCHE":
            return "background-color: #e8f3fb; color: #0f6095; font-weight: 700;"
        if valor == "DESCANSO":
            return "background-color: #fff6d6; color: #8a6610; font-weight: 700;"
        return ""

    return (
        tabla.style
        .map(color_celda)
        .set_properties(
            subset=["GRUPO", "PUESTO", "# OF."],
            **{"font-weight": "700", "background-color": "#f4f6f9"},
        )
        .set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#7c2229"),
                        ("color", "#ffffff"),
                        ("font-weight", "800"),
                    ],
                }
            ]
        )
    )


st.set_page_config(
    page_title="Seguridad Total | Roles",
    page_icon="🛡️",
    layout="wide",
)

aplicar_estilos_marca()
render_hero()

st.caption("Basado en protocolos de prevención, cobertura y supervisión operativa.")

cliente = st.text_input("Cliente o servicio", value="Denso Silao")

tipo_turnos = st.selectbox(
    "Esquema de cobertura",
    ["Día y noche", "Solo día", "Solo noche"],
)

dia_fijo = False

if tipo_turnos in ("Día y noche", "Solo día"):
    dia_fijo = st.checkbox(
        "Personal de día fijo",
        value=True,
        help=(
            "Cubre lunes a viernes con oficiales fijos 5x2. "
            "Los fines de semana o aumentos se acomodan aparte."
        ),
    )

guardias_dia = [0] * 7
guardias_noche = [0] * 7

if tipo_turnos in ("Día y noche", "Solo día"):
    st.subheader("Cobertura requerida | Turno de día")
    columnas_dia = st.columns(7)
    valores_dia = [5, 5, 5, 5, 6, 5, 5]

    for indice, dia in enumerate(DIAS):
        with columnas_dia[indice]:
            cantidad = st.number_input(
                dia,
                min_value=0,
                value=valores_dia[indice],
                step=1,
                key=f"dia_{indice}",
            )
            guardias_dia[indice] = int(cantidad)

if tipo_turnos in ("Día y noche", "Solo noche"):
    st.subheader("Cobertura requerida | Turno de noche")
    columnas_noche = st.columns(7)

    for indice, dia in enumerate(DIAS):
        with columnas_noche[indice]:
            cantidad = st.number_input(
                dia,
                min_value=0,
                value=1,
                step=1,
                key=f"noche_{indice}",
            )
            guardias_noche[indice] = int(cantidad)

maximo_posiciones_12x24 = min(min(guardias_dia), min(guardias_noche))
posiciones_12x24 = 0

st.subheader("Configuración de la rotación")
columna_1, columna_2 = st.columns(2)

with columna_1:
    horas_turno = st.number_input(
        "Horas por turno",
        min_value=1,
        max_value=24,
        value=12,
    )

with columna_2:
    horas_objetivo = st.number_input(
        "Horas objetivo semanales",
        min_value=1,
        max_value=168,
        value=48,
    )

max_turnos_guardia = 5 if int(horas_turno) <= 12 else max(
    1,
    min(7, int(horas_objetivo) // int(horas_turno)),
)

try:
    roles = generar_roles(
        guardias_dia=guardias_dia,
        guardias_noche=guardias_noche,
        posiciones_12x24=int(posiciones_12x24),
        max_turnos_guardia=int(max_turnos_guardia),
        dia_fijo=dia_fijo,
    )
except ValueError as error:
    st.error(str(error))
    st.stop()

dia_s1, noche_s1 = contar_cobertura(roles, 0)
dia_s2, noche_s2 = contar_cobertura(roles, 7)

cobertura_correcta = (
    dia_s1 == guardias_dia
    and dia_s2 == guardias_dia
    and noche_s1 == guardias_noche
    and noche_s2 == guardias_noche
)

total_turnos = sum(guardias_dia) + sum(guardias_noche)

st.divider()
st.subheader("Panel de resultado")
resultado_1, resultado_2, resultado_3, resultado_4 = st.columns(4)
resultado_1.metric("Turnos semanales", total_turnos)
resultado_2.metric("Guardias generados", len(roles))
resultado_3.metric("Cobertura día", sum(guardias_dia))
resultado_4.metric("Cobertura noche", sum(guardias_noche))

if cobertura_correcta:
    st.success("Cobertura validada: las dos semanas coinciden con lo solicitado.")
else:
    st.error("La cobertura generada no coincide. Revisa los parámetros.")

st.subheader("Vista previa del rol quincenal")
tabla_previa = crear_tabla_previa(
    roles,
    int(horas_turno),
    int(horas_objetivo),
)
st.dataframe(pintar_turnos(tabla_previa), use_container_width=True, hide_index=True)

excel = generar_excel(
    cliente=cliente,
    tipo_turnos=tipo_turnos,
    guardias_dia=guardias_dia,
    guardias_noche=guardias_noche,
    roles=roles,
    horas_turno=int(horas_turno),
    horas_objetivo=int(horas_objetivo),
)

nombre_archivo = (
    cliente.lower()
    .strip()
    .replace(" ", "_")
    .replace("/", "_")
    .replace("\\", "_")
)

st.download_button(
    label="Descargar rol en Excel",
    data=excel,
    file_name=f"rol_{nombre_archivo}.xlsx",
    mime=(
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    ),
    use_container_width=True,
)
